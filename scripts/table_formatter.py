#!/usr/bin/env python3
"""
Table Formatting Utilities

Provides functions to ensure proper table formatting:
- Word wrapping for long text
- Proper column widths
- No text overflow
- Headers fit within cells
"""

import pandas as pd
from typing import Dict, Optional
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def format_dataframe_for_display(df: pd.DataFrame, 
                                 max_column_width: int = 50,
                                 wrap_text: bool = True) -> pd.DataFrame:
    """
    Format DataFrame for display to prevent overflow.
    
    Args:
        df: Input DataFrame
        max_column_width: Maximum characters per column
        wrap_text: Whether to wrap long text
        
    Returns:
        Formatted DataFrame with truncated/wrapped text
    """
    if df.empty:
        return df
    
    formatted_df = df.copy()
    
    for col in formatted_df.columns:
        if formatted_df[col].dtype == 'object':  # String columns
            # Truncate very long text
            formatted_df[col] = formatted_df[col].astype(str).apply(
                lambda x: x[:max_column_width] + '...' if len(str(x)) > max_column_width else str(x)
            )
    
    return formatted_df


def format_excel_column_widths(workbook_path: str, 
                               sheet_name: Optional[str] = None,
                               column_widths: Optional[Dict[str, int]] = None,
                               wrap_text: bool = True):
    """
    Format Excel file column widths and enable word wrapping.
    
    Args:
        workbook_path: Path to Excel file
        sheet_name: Sheet name (None for first sheet)
        column_widths: Dict of column letters to widths (e.g., {'A': 15, 'B': 20})
        wrap_text: Enable word wrapping in cells
    """
    try:
        wb = load_workbook(workbook_path)
        ws = wb.active if sheet_name is None else wb[sheet_name]
        
        # Set column widths
        if column_widths:
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
        else:
            # Auto-size columns based on content
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Set width (with some padding, max 50)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Enable word wrapping for all cells
        if wrap_text:
            alignment = Alignment(wrap_text=True, vertical='top')
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = alignment
        
        # Format header row
        if ws.max_row > 0:
            header_font = Font(bold=True, size=11)
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            # Set header row height for wrapped text
            ws.row_dimensions[1].height = 30
        
        wb.save(workbook_path)
        
    except Exception as e:
        print(f"Warning: Could not format Excel file {workbook_path}: {e}")


def format_csv_for_display(df: pd.DataFrame, 
                           max_column_width: int = 50) -> pd.DataFrame:
    """
    Format DataFrame before CSV export to prevent overflow.
    
    Args:
        df: Input DataFrame
        max_column_width: Maximum characters per column
        
    Returns:
        Formatted DataFrame
    """
    if df.empty:
        return df
    
    formatted_df = df.copy()
    
    for col in formatted_df.columns:
        if formatted_df[col].dtype == 'object':  # String columns
            # Truncate long text with ellipsis
            formatted_df[col] = formatted_df[col].astype(str).apply(
                lambda x: (x[:max_column_width] + '...') if len(str(x)) > max_column_width else str(x)
            )
    
    return formatted_df


def get_smart_column_widths(df: pd.DataFrame, 
                            min_width: int = 10,
                            max_width: int = 50) -> Dict[str, int]:
    """
    Calculate smart column widths based on content.
    
    Args:
        df: Input DataFrame
        min_width: Minimum column width
        max_width: Maximum column width
        
    Returns:
        Dict of column names to widths
    """
    widths = {}
    
    for col in df.columns:
        # Start with header width
        header_width = len(str(col))
        
        # Check content widths
        content_width = 0
        if not df.empty:
            sample_values = df[col].astype(str).head(10)
            if len(sample_values) > 0:
                content_width = sample_values.str.len().max()
        
        # Use the larger of header or content, with limits
        width = max(header_width, content_width, min_width)
        width = min(width, max_width)
        
        widths[col] = width
    
    return widths



