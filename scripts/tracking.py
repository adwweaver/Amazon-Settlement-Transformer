#!/usr/bin/env python3
"""
Amazon Settlement ETL Pipeline - Entry Tracking Module

This module handles tracking of settlement file processing and Zoho Books entry status.
Maintains Entry_Status.csv in the SharePoint folder to track the complete lifecycle.

Author: ETL Pipeline
Date: October 2025
"""

import logging
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Optional
import os


class EntryTracker:
    """
    Tracks settlement processing and Zoho Books entry status.
    
    Valid Status Values:
    - Processed - Pending Approval: File processed, awaiting review
    - Requires Review - Out of Balance: Processing complete but mathematical issues found
    - Approved - Ready for Zoho: Reviewed and approved for entry
    - Uploaded to Zoho - Complete: Successfully entered into Zoho Books
    - On Hold - Issue Found: Processing paused due to data issues
    """
    
    # Status constants
    STATUS_PENDING_APPROVAL = "Processed - Pending Approval"
    STATUS_OUT_OF_BALANCE = "Requires Review - Out of Balance"
    STATUS_APPROVED = "Approved - Ready for Zoho"
    STATUS_COMPLETE = "Uploaded to Zoho - Complete"
    STATUS_ON_HOLD = "On Hold - Issue Found"
    
    VALID_STATUSES = [
        STATUS_PENDING_APPROVAL,
        STATUS_OUT_OF_BALANCE,
        STATUS_APPROVED,
        STATUS_COMPLETE,
        STATUS_ON_HOLD
    ]
    
    def __init__(self, tracking_file: str = None):
        """
        Initialize the EntryTracker.
        
        Args:
            tracking_file: Path to Entry_Status.csv. If None, uses default SharePoint location.
        """
        self.logger = logging.getLogger(__name__)
        
        # Default to SharePoint location if not specified
        if tracking_file is None:
            sharepoint_base = Path(os.path.expanduser(
                r"~\Touchstone Brands\BrackishCo - Documents\Sharepoint_Public\Amazon-ETL"
            ))
            self.tracking_file = sharepoint_base / "Entry_Status.xlsx"
        else:
            self.tracking_file = Path(tracking_file)
        
        # Ensure file exists with headers
        if not self.tracking_file.exists():
            self._create_tracking_file()
    
    def _create_tracking_file(self):
        """Create tracking file with headers if it doesn't exist."""
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        
        headers = [
            'Settlement_ID',
            'Deposit_Date',
            'Deposit_Amount',
            'Date_Processed',
            'Processed_By',
            'Date_Approved',
            'Approved_By',
            'Date_Entered_Zoho',
            'Entered_By',
            'Status',
            'Notes'
        ]
        
        # Create Excel file with data validation
        df = pd.DataFrame(columns=headers)
        df.to_excel(self.tracking_file, index=False, sheet_name='Entry Status', engine='openpyxl')
        
        # Add dropdown validation for Status column
        from openpyxl import load_workbook
        wb = load_workbook(self.tracking_file)
        ws = wb.active
        
        status_options = (
            '"Processed - Pending Approval,'
            'Requires Review - Out of Balance,'
            'Approved - Ready for Zoho,'
            'Uploaded to Zoho - Complete,'
            'On Hold - Issue Found"'
        )
        
        dv = DataValidation(
            type="list",
            formula1=status_options,
            allow_blank=True
        )
        dv.error = 'Invalid status selected'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = 'Please select a status from the dropdown'
        dv.promptTitle = 'Status Selection'
        dv.add('J2:J1000')
        ws.add_data_validation(dv)
        
        # Professional formatting for headers
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Header row formatting: Bold, larger font, background color, wrapped text
        header_font = Font(bold=True, size=11, name='Calibri')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Professional blue
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # Make text white for better contrast on blue background
            cell.font = Font(bold=True, size=11, name='Calibri', color='FFFFFF')
        
        # Increase header row height for wrapped text
        ws.row_dimensions[1].height = 30
        
        # Set column widths for better readability
        column_widths = {
            'A': 15,  # Settlement_ID
            'B': 17,  # Deposit_Date (+5)
            'C': 20,  # Deposit_Amount (+5)
            'D': 20,  # Date_Processed
            'E': 15,  # Processed_By
            'F': 20,  # Date_Approved
            'G': 15,  # Approved_By
            'H': 20,  # Date_Entered_Zoho
            'I': 15,  # Entered_By
            'J': 30,  # Status (wider for dropdown text)
            'K': 40   # Notes (widest for comments)
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze the header row so it stays visible when scrolling
        ws.freeze_panes = 'A2'
        
        wb.save(self.tracking_file)
        self.logger.info(f"Created tracking file: {self.tracking_file}")
    
    def record_processing(self, settlement_id: str, deposit_date: str = "", deposit_amount: float = 0.0, 
                         processed_by: str = "ETL Pipeline", notes: str = ""):
        """
        Record that a settlement file has been processed.
        
        Args:
            settlement_id: The settlement ID that was processed
            deposit_date: The deposit date for this settlement (YYYY-MM-DD)
            deposit_amount: The deposit amount
            processed_by: Who/what processed it (default: "ETL Pipeline")
            notes: Optional notes about the processing
        """
        try:
            # Load existing tracking data
            df = pd.read_excel(self.tracking_file, engine='openpyxl')
            
            # Check if this settlement already exists
            existing = df[df['Settlement_ID'] == settlement_id]
            
            if len(existing) > 0:
                # Update existing record
                idx = existing.index[0]
                df.loc[idx, 'Deposit_Date'] = deposit_date
                df.loc[idx, 'Deposit_Amount'] = deposit_amount
                df.loc[idx, 'Date_Processed'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                df.loc[idx, 'Processed_By'] = processed_by
                df.loc[idx, 'Status'] = self.STATUS_PENDING_APPROVAL
                if notes:
                    df.loc[idx, 'Notes'] = notes
                self.logger.info(f"Updated tracking for settlement {settlement_id}")
            else:
                # Add new record
                new_row = {
                    'Settlement_ID': settlement_id,
                    'Deposit_Date': deposit_date,
                    'Deposit_Amount': deposit_amount,
                    'Date_Processed': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Processed_By': processed_by,
                    'Date_Approved': '',
                    'Approved_By': '',
                    'Date_Entered_Zoho': '',
                    'Entered_By': '',
                    'Status': 'Processed - Pending Approval',
                    'Notes': notes
                }
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                self.logger.info(f"Added tracking for settlement {settlement_id}")
            
            # Save updated tracking file
            df.to_excel(self.tracking_file, index=False, sheet_name='Entry Status', engine='openpyxl')
            
        except Exception as e:
            self.logger.error(f"Error recording processing for {settlement_id}: {e}")
    
    def mark_approved(self, settlement_id: str, approved_by: str, notes: str = ""):
        """
        Mark a settlement as approved and ready for Zoho entry.
        
        Args:
            settlement_id: The settlement ID to approve
            approved_by: Who approved it
            notes: Optional approval notes
        """
        try:
            df = pd.read_excel(self.tracking_file, engine='openpyxl')
            existing = df[df['Settlement_ID'] == settlement_id]
            
            if len(existing) == 0:
                self.logger.warning(f"Settlement {settlement_id} not found in tracking")
                return
            
            idx = existing.index[0]
            df.loc[idx, 'Date_Approved'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            df.loc[idx, 'Approved_By'] = approved_by
            df.loc[idx, 'Status'] = self.STATUS_APPROVED
            if notes:
                current_notes = df.loc[idx, 'Notes']
                df.loc[idx, 'Notes'] = f"{current_notes}; {notes}" if current_notes else notes
            
            df.to_excel(self.tracking_file, index=False, sheet_name='Entry Status', engine='openpyxl')
            self.logger.info(f"Marked settlement {settlement_id} as approved")
            
        except Exception as e:
            self.logger.error(f"Error marking approval for {settlement_id}: {e}")
    
    def mark_entered_zoho(self, settlement_id: str, entered_by: str, notes: str = ""):
        """
        Mark a settlement as entered into Zoho Books.
        
        Args:
            settlement_id: The settlement ID that was entered
            entered_by: Who entered it (or "Zoho API" for automated)
            notes: Optional entry notes
        """
        try:
            df = pd.read_excel(self.tracking_file, engine='openpyxl')
            existing = df[df['Settlement_ID'] == settlement_id]
            
            if len(existing) == 0:
                self.logger.warning(f"Settlement {settlement_id} not found in tracking")
                return
            
            idx = existing.index[0]
            df.loc[idx, 'Date_Entered_Zoho'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            df.loc[idx, 'Entered_By'] = entered_by
            df.loc[idx, 'Status'] = self.STATUS_COMPLETE
            if notes:
                current_notes = df.loc[idx, 'Notes']
                df.loc[idx, 'Notes'] = f"{current_notes}; {notes}" if current_notes else notes
            
            df.to_excel(self.tracking_file, index=False, sheet_name='Entry Status', engine='openpyxl')
            self.logger.info(f"Marked settlement {settlement_id} as entered in Zoho")
            
        except Exception as e:
            self.logger.error(f"Error marking Zoho entry for {settlement_id}: {e}")
    
    def get_status(self, settlement_id: str) -> Optional[dict]:
        """
        Get the current status of a settlement.
        
        Args:
            settlement_id: The settlement ID to look up
            
        Returns:
            Dictionary with status info, or None if not found
        """
        try:
            df = pd.read_excel(self.tracking_file, engine='openpyxl')
            existing = df[df['Settlement_ID'] == settlement_id]
            
            if len(existing) == 0:
                return None
            
            return existing.iloc[0].to_dict()
            
        except Exception as e:
            self.logger.error(f"Error getting status for {settlement_id}: {e}")
            return None
    
    def get_pending_approval(self) -> pd.DataFrame:
        """Get all settlements pending approval."""
        try:
            df = pd.read_excel(self.tracking_file, engine='openpyxl')
            return df[df['Status'] == self.STATUS_PENDING_APPROVAL]
        except Exception as e:
            self.logger.error(f"Error getting pending approvals: {e}")
            return pd.DataFrame()
    
    def get_pending_zoho_entry(self) -> pd.DataFrame:
        """Get all settlements approved but not yet entered in Zoho."""
        try:
            df = pd.read_excel(self.tracking_file, engine='openpyxl')
            return df[df['Status'] == self.STATUS_APPROVED]
        except Exception as e:
            self.logger.error(f"Error getting pending Zoho entries: {e}")
            return pd.DataFrame()
    
    def record_settlement_history(self, settlement_data: dict) -> bool:
        """
        Record settlement details to historical log (append-only CSV).
        Used for trending, analysis, and financial reporting.
        
        Args:
            settlement_data: Dictionary with keys:
                - settlement_id
                - deposit_date
                - date_from
                - date_to
                - bank_deposit_amount
                - total_records
                - journal_line_count
                - invoice_line_count
                - tax_line_count
                - gl_account_totals (dict of GL account: amount)
                - date_processed
                
        Returns:
            True if successful, False otherwise
        """
        try:
            # Save to SharePoint location
            from paths import get_settlement_history_path
            history_file = get_settlement_history_path()
            history_file.parent.mkdir(exist_ok=True)
            
            # Flatten GL account totals into separate columns
            row_data = {
                'settlement_id': settlement_data.get('settlement_id'),
                'deposit_date': settlement_data.get('deposit_date'),
                'date_from': settlement_data.get('date_from'),
                'date_to': settlement_data.get('date_to'),
                'bank_deposit_amount': settlement_data.get('bank_deposit_amount', 0),
                'total_records': settlement_data.get('total_records', 0),
                'journal_line_count': settlement_data.get('journal_line_count', 0),
                'invoice_line_count': settlement_data.get('invoice_line_count', 0),
                'tax_line_count': settlement_data.get('tax_line_count', 0),
                'date_processed': settlement_data.get('date_processed', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                # Zoho Books sync tracking
                'zoho_synced': False,
                'zoho_journal_id': None,
                'zoho_sync_date': None,
                'zoho_sync_status': 'pending'
            }
            
            # Add GL account totals as individual columns
            gl_totals = settlement_data.get('gl_account_totals', {})
            for gl_account, amount in gl_totals.items():
                # Clean GL account name for column name
                col_name = f"gl_{gl_account.replace(' ', '_').replace('.', '').replace('-', '_')}"
                # Ensure amount is a valid number
                try:
                    row_data[col_name] = float(amount) if pd.notna(amount) else 0.0
                except (ValueError, TypeError):
                    row_data[col_name] = 0.0
            
            # Load existing history or create new
            if history_file.exists():
                try:
                    df = pd.read_csv(history_file)
                    # Check if settlement already exists
                    if settlement_data.get('settlement_id') in df['settlement_id'].values:
                        # Update existing record
                        mask = df['settlement_id'] == settlement_data.get('settlement_id')
                        for col, val in row_data.items():
                            if col in df.columns:
                                df.loc[mask, col] = val
                            else:
                                df[col] = None
                                df.loc[mask, col] = val
                    else:
                        # Append new record
                        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)
                except Exception as e:
                    self.logger.warning(f"Could not load existing history, creating new: {e}")
                    df = pd.DataFrame([row_data])
            else:
                df = pd.DataFrame([row_data])
            
            # Save to CSV
            df.to_csv(history_file, index=False)
            self.logger.info(f"Settlement history recorded for {settlement_data.get('settlement_id')}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error recording settlement history: {e}")
            return False


if __name__ == "__main__":
    # Example usage
    tracker = EntryTracker()
    
    # Example: Record processing
    tracker.record_processing("12345678901", notes="Test settlement")
    
    # Example: Get status
    status = tracker.get_status("12345678901")
    print(f"Status: {status}")
    
    # Example: Mark approved
    tracker.mark_approved("12345678901", "John Doe", "Amounts verified")
    
    # Example: Mark entered in Zoho
    tracker.mark_entered_zoho("12345678901", "John Doe", "All entries posted")
